from __future__ import annotations

import asyncio
import csv
from io import StringIO
import os
from pathlib import Path
from typing import Any, List

from backend.file_processors.parsers.base_parser import BaseParser, ParsedContent
from backend.file_processors.parsers.text_readers import read_text_file_with_fallback


class ExcelParser(BaseParser):
    # 常见表头关键词：命中后优先按“有表头”处理，避免把规范表头误判成第一行数据。
    _COMMON_HEADER_TOKENS = {
        "名称", "说明", "描述", "标题", "内容", "类型", "状态", "备注", "金额", "数量", "日期",
        "name", "title", "description", "content", "type", "status", "remark", "remarks", "amount", "count", "date", "id",
    }

    async def parse(self, file_path: str) -> ParsedContent:
        file_extension = Path(file_path).suffix.lower()
        if file_extension == ".xlsx":
            return await asyncio.to_thread(self._parse_xlsx_sync, file_path)

        return await asyncio.to_thread(self._parse_legacy_excel_sync, file_path)

    def _parse_xlsx_sync(self, file_path: str) -> ParsedContent:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is not installed. Please install it: pip install openpyxl") from exc

        try:
            all_text: List[str] = []
            tables_data = []
            cell_profile = self._get_text_cleaning_profile(".xlsx")
            # 公式工作簿保留公式、批注和超链接；值工作簿优先提供缓存显示值。
            formula_workbook = load_workbook(file_path, read_only=False, data_only=False)
            value_workbook = load_workbook(file_path, read_only=False, data_only=True)

            try:
                for worksheet in formula_workbook.worksheets:
                    cleaned_sheet_name = self._clean_text_value(worksheet.title, cell_profile)
                    value_worksheet = value_workbook[worksheet.title] if worksheet.title in value_workbook.sheetnames else None
                    sheet_rows: list[list[str]] = []
                    max_row = max(int(getattr(worksheet, "max_row", 0) or 0), int(getattr(value_worksheet, "max_row", 0) or 0))
                    max_column = max(int(getattr(worksheet, "max_column", 0) or 0), int(getattr(value_worksheet, "max_column", 0) or 0))

                    for row_index in range(1, max_row + 1):
                        cleaned_row = []
                        for column_index in range(1, max_column + 1):
                            formula_cell = worksheet.cell(row=row_index, column=column_index)
                            value_cell = value_worksheet.cell(row=row_index, column=column_index) if value_worksheet is not None else None
                            cleaned_row.append(self._clean_text_value(self._format_excel_cell(formula_cell, value_cell), cell_profile))
                        if any(cleaned_row):
                            sheet_rows.append(cleaned_row)

                    if not sheet_rows:
                        continue

                    columns, rows, has_header = self._normalize_table_rows(sheet_rows)
                    sheet_lines = [f"[工作表] {cleaned_sheet_name}"]
                    for row_index, cleaned_row in enumerate(rows, start=1):
                        row_values = self._build_row_value_pairs(columns, cleaned_row)
                        if row_values:
                            sheet_lines.append(f"第{row_index}行 | " + " | ".join(row_values))

                    all_text.append("\n".join(sheet_lines))
                    tables_data.append(
                        {
                            "sheet_name": cleaned_sheet_name,
                            "rows": rows,
                            "columns": columns,
                            "has_header": has_header,
                        }
                    )
            finally:
                formula_workbook.close()
                value_workbook.close()

            full_text = "\n\n".join(all_text)
            metadata = {
                "sheet_count": len(tables_data),
                "sheet_names": [table["sheet_name"] for table in tables_data],
                "has_header": all(bool(table.get("has_header", True)) for table in tables_data) if tables_data else True,
                "parser": "openpyxl_structured",
                "has_text": bool(full_text),
                "empty_content": not bool(full_text),
            }

            return ParsedContent(text=full_text, metadata=metadata, tables=tables_data)

        except Exception as exc:
            self.logger.error(f"Failed to parse Excel: {str(exc)}")
            raise

    def _parse_legacy_excel_sync(self, file_path: str) -> ParsedContent:
        try:
            import pandas as pd
        except ImportError as exc:
            raise ImportError("pandas is not installed. Please install it: pip install pandas xlrd") from exc

        try:
            excel_file = pd.ExcelFile(file_path)
            all_text: List[str] = []
            tables_data = []
            cell_profile = self._get_text_cleaning_profile(Path(file_path).suffix.lower())

            for sheet_name in excel_file.sheet_names:
                dataframe = pd.read_excel(file_path, sheet_name=sheet_name).fillna("")
                cleaned_sheet_name = self._clean_text_value(sheet_name, cell_profile)
                raw_rows = [[
                    self._clean_text_value(column, cell_profile)
                    for column in dataframe.columns.tolist()
                ]]
                raw_rows.extend(
                    [self._clean_text_value(value, cell_profile) for value in row]
                    for row in dataframe.itertuples(index=False, name=None)
                )
                columns, rows, has_header = self._normalize_table_rows(raw_rows)
                sheet_lines = [f"[工作表] {cleaned_sheet_name}"]
                for row_index, cleaned_row in enumerate(rows, start=1):
                    row_values = self._build_row_value_pairs(columns, cleaned_row)
                    if row_values:
                        sheet_lines.append(f"第{row_index}行 | " + " | ".join(row_values))

                all_text.append("\n".join(sheet_lines))
                tables_data.append(
                    {
                        "sheet_name": cleaned_sheet_name,
                        "rows": rows,
                        "columns": columns,
                        "has_header": has_header,
                    }
                )

            full_text = "\n\n".join(all_text)
            metadata = {
                "sheet_count": len(excel_file.sheet_names),
                "sheet_names": excel_file.sheet_names,
                "has_header": all(bool(table.get("has_header", True)) for table in tables_data) if tables_data else True,
                "parser": "pandas_xls",
                "has_text": bool(full_text),
                "empty_content": not bool(full_text),
            }

            return ParsedContent(text=full_text, metadata=metadata, tables=tables_data)

        except Exception as exc:
            self.logger.error(f"Failed to parse legacy Excel: {str(exc)}")
            raise

    @classmethod
    def _normalize_table_rows(cls, sheet_rows: list[list[str]]) -> tuple[list[str], list[list[str]], bool]:
        if not sheet_rows:
            return [], [], True

        max_width = max(len(row) for row in sheet_rows)
        normalized_rows = [row[:max_width] + [""] * max(0, max_width - len(row)) for row in sheet_rows]
        header_row = normalized_rows[0]
        has_header = cls._looks_like_header_row(header_row, normalized_rows[1:4])

        if has_header:
            columns = [
                (header_row[index] if index < len(header_row) else "") or f"column_{index + 1}"
                for index in range(max_width)
            ]
            rows = normalized_rows[1:]
            return columns, rows, True

        columns = [f"column_{index + 1}" for index in range(max_width)]
        return columns, normalized_rows, False

    @classmethod
    def _looks_like_header_row(cls, header_row: list[str], sample_rows: list[list[str]]) -> bool:
        normalized_cells = [str(cell or "").strip() for cell in header_row if str(cell or "").strip()]
        if not normalized_cells:
            return False

        normalized_lower = [cell.lower() for cell in normalized_cells]
        if any(cell in cls._COMMON_HEADER_TOKENS for cell in normalized_lower):
            return True

        if all(cls._looks_like_data_value(cell) for cell in normalized_cells):
            return False

        if sample_rows:
            sample_values = [str(cell or "").strip() for row in sample_rows for cell in row if str(cell or "").strip()]
            if sample_values and any(cls._looks_like_data_value(cell) for cell in sample_values) and not any(
                cls._looks_like_data_value(cell) for cell in normalized_cells
            ):
                return True

        # 回退判定必须保守：若没有明显表头信号，就把首行视为数据，避免误吃掉真实首行。
        return False

    @staticmethod
    def _looks_like_data_value(value: str) -> bool:
        normalized_value = str(value or "").strip()
        if not normalized_value:
            return False
        if normalized_value.startswith("="):
            return True
        if normalized_value.lower().startswith(("http://", "https://", "www.")):
            return True
        if normalized_value.replace(".", "", 1).replace(",", "", 1).isdigit():
            return True
        return False

    @staticmethod
    def _build_row_value_pairs(columns: list[str], row: list[str]) -> list[str]:
        return [
            f"{column}: {value}"
            for column, value in zip(columns, row)
            if value
        ]

    @staticmethod
    def _format_excel_cell(cell, value_cell: Any | None = None) -> str:
        if cell is None:
            return ""

        value = getattr(cell, "value", None)
        cached_value = getattr(value_cell, "value", None) if value_cell is not None else None
        result = ""
        if getattr(cell, "data_type", None) == "f":
            formula_text = str(value)
            formula_text = formula_text if formula_text.startswith("=") else f"={formula_text}"
            if cached_value not in (None, ""):
                result = f"{cached_value} [formula: {formula_text}]"
            else:
                result = formula_text
        elif cached_value not in (None, ""):
            result = str(cached_value)
        elif value is not None:
            result = str(value)

        if not result:
            return ""

        hyperlink = getattr(getattr(cell, "hyperlink", None), "target", None)
        comment = getattr(getattr(cell, "comment", None), "text", None)
        if hyperlink:
            result = f"{result} ({hyperlink})"
        if comment:
            result = f"{result} [comment: {comment}]"
        return result

    def supports(self, file_extension: str) -> bool:
        return file_extension.lower() in [".xlsx", ".xls"]

    def get_supported_extensions(self) -> List[str]:
        return [".xlsx", ".xls"]


class TabularParser(BaseParser):
    """CSV/TSV 独立解析器。

    这里不再把 `csv/tsv` 混入 ExcelParser，避免扩展名 -> FileType -> Parser 漂移。
    """

    async def parse(self, file_path: str) -> ParsedContent:
        file_extension = Path(file_path).suffix.lower()
        delimiter = "\t" if file_extension == ".tsv" else ","
        text, used_encoding = self._read_text_file(file_path)

        table_name = Path(file_path).stem or file_extension.lstrip(".") or "table"
        cell_profile = self._get_text_cleaning_profile(file_extension)
        # StringIO + newline="" 可以保留 CSV/TSV 引号字段内的真实换行。
        raw_rows = list(csv.reader(StringIO(text, newline=""), delimiter=delimiter))
        cleaned_rows = [
            [self._clean_text_value(cell, cell_profile) for cell in row]
            for row in raw_rows
            if any(str(cell or "").strip() for cell in row)
        ]

        if cleaned_rows:
            columns, rows, has_header = ExcelParser._normalize_table_rows(cleaned_rows)
        else:
            columns = []
            rows = []
            has_header = True

        preview_lines = [f"[表格] {table_name}"]
        if columns:
            preview_lines.append("列: " + " | ".join(columns))
        for row_index, row in enumerate(rows, start=1):
            row_values = ExcelParser._build_row_value_pairs(columns, row)
            if row_values:
                preview_lines.append(f"第{row_index}行 | " + " | ".join(row_values))

        metadata = {
            "encoding": used_encoding,
            "delimiter": delimiter,
            "delimiter_name": "tab" if delimiter == "\t" else "comma",
            "table_format": file_extension.lstrip("."),
            "table_count": 1,
            "row_count": len(rows),
            "column_count": len(columns),
            "has_header": has_header,
            "parser": "tabular",
        }
        tables = [
            {
                "sheet_name": table_name,
                "columns": columns,
                "rows": rows,
                "has_header": has_header,
                "table_type": file_extension.lstrip("."),
            }
        ]
        return ParsedContent(text="\n".join(preview_lines), metadata=metadata, tables=tables)

    def _read_text_file(self, file_path: str) -> tuple[str, str]:
        return read_text_file_with_fallback(file_path)

    def supports(self, file_extension: str) -> bool:
        return file_extension.lower() in [".csv", ".tsv"]

    def get_supported_extensions(self) -> List[str]:
        return [".csv", ".tsv"]


class TextParser(BaseParser):
    async def parse(self, file_path: str) -> ParsedContent:
        try:
            text, used_encoding = read_text_file_with_fallback(file_path)

            file_ext = os.path.splitext(file_path)[1].lower()
            metadata = {
                "encoding": used_encoding,
                "file_extension": file_ext,
                "line_count": len(text.split("\n")),
                "char_count": len(text),
                "parser": "text",
                "has_text": bool(text),
                "empty_content": not bool(text),
            }

            return ParsedContent(text=text, metadata=metadata)

        except Exception as exc:
            self.logger.error(f"Failed to parse text file: {str(exc)}")
            raise

    def supports(self, file_extension: str) -> bool:
        return file_extension.lower() in [
            ".txt",
            ".md",
            ".markdown",
            ".json",
            ".xml",
            ".svg",
            ".log",
            ".rst",
            ".py",
            ".js",
            ".ts",
            ".tsx",
            ".jsx",
            ".java",
            ".cpp",
            ".c",
            ".h",
            ".hpp",
            ".go",
            ".rs",
            ".sh",
            ".bash",
            ".bat",
            ".ps1",
            ".sql",
            ".css",
            ".scss",
            ".less",
            ".vue",
            ".yaml",
            ".yml",
            ".toml",
            ".ini",
            ".conf",
            ".cfg",
            ".env",
            ".properties",
            ".gitignore",
        ]

    def get_supported_extensions(self) -> List[str]:
        return [
            ".txt",
            ".md",
            ".markdown",
            ".json",
            ".xml",
            ".svg",
            ".log",
            ".rst",
            ".py",
            ".js",
            ".ts",
            ".tsx",
            ".jsx",
            ".java",
            ".cpp",
            ".c",
            ".h",
            ".hpp",
            ".go",
            ".rs",
            ".sh",
            ".bash",
            ".bat",
            ".ps1",
            ".sql",
            ".css",
            ".scss",
            ".less",
            ".vue",
            ".yaml",
            ".yml",
            ".toml",
            ".ini",
            ".conf",
            ".cfg",
            ".env",
            ".properties",
            ".gitignore",
        ]
